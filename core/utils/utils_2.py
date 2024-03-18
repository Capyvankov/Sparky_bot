import openpyxl

book = openpyxl.load_workbook(filename='schedule.xlsx')
ws = book.active

def get_column_for_class(num_and_letter: str) -> int:
    try:
        for i in range(1, ws.max_row):
            current_cell = ws.cell(row=4, column=i)
            if num_and_letter.lower() == current_cell.value:
                return current_cell.column
        print('Такого класса не существует')
    except:
        print("Фунцкция get_column_for_class не смогла найти искомое")

def get_day_of_week_row(week_day: str) -> int:
    try:
        for i in range(1, ws.max_column):
            current_cell = ws.cell(row=i, column=1)
            if week_day.lower() == current_cell.value:
                # print('Найденный столбик: ', current_cell.column, '\nАдрес ячейки: ', current_cell.coordinate)
                return current_cell.row
        print('Такого дня не существует')
    except:
        print("Фунцкция get_day_of_week_row не смогла найти искомое")

def get_schedule(week_day: str, num_and_letter: str) -> list:
    sch_for_day = []
    sch_room = []
    try:
        current_row = get_day_of_week_row(week_day)
        current_column = get_column_for_class(num_and_letter)
        for i in range(7):
            current_cell = ws.cell(row=current_row+i, column=current_column)
            current_cell_room = ws.cell(row=current_row + i, column=current_column+1)
            if current_cell.value is None:
                continue
            if current_cell_room.value is None:
                continue
            sch_for_day.append(current_cell.value)
            sch_room.append(current_cell_room.value)
            # print('Найденный столбик: ', current_cell.column, '\nАдрес ячейки: ', current_cell.coordinate)
        return sch_for_day, sch_room

    except:
        print("Фунцкция get_schedule не смогла найти искомое")

# def get_schedule_for_all_week(num_and_letter: str) -> list:
#     try:
#         current_column = get_column_for_class(num_and_letter)
#         sch_for_day = []
#         for i in range(35):
#             current_cell = ws.cell(row=5+i, column=current_column)
#             if current_cell.value is None:
#                 continue
#             else:
#                 sch_for_day.append(current_cell.value)
#             # print('Найденный столбик: ', current_cell.column, '\nАдрес ячейки: ', current_cell.coordinate)
#         return sch_for_day
#
#     except:
#         print("Фунцкция get_schedule_for_all_week не смогла найти искомое")
#
# print(get_schedule('Понедельник', ''))























# example_list = []
#
# def sch_save_5a():
#     for row in ws.iter_rows(min_row=4, min_col=3, max_col=3, max_row=46, values_only=True):
#         example_list.append(row)
#
#     with open("5a.txt", "w", encoding='UTF8') as file:
#         for elem in example_list:
#             file.write(str(elem) + '\n')
#
#     with open("5a.txt", "r") as file:
#         print(file)

